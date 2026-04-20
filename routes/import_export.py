import logging
from flask import Blueprint, jsonify, request
from datetime import datetime
from io import BytesIO
from models import (get_db, validate_date, validate_number, validate_pct,
                    validate_string, validate_isin)
from auth import login_required, csrf_protect

MAX_IMPORT_ROWS = 10000
MAX_NOTE_LENGTH = 2000

logger = logging.getLogger('financy')

import_export_bp = Blueprint('import_export', __name__)


@import_export_bp.route('/api/import', methods=['POST'])
@login_required
@csrf_protect
def import_xlsx():
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Seuls les fichiers .xlsx sont acceptés'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        imported = 0
        skipped = 0

        def _parse_date(val):
            if isinstance(val, datetime):
                return val.strftime('%Y-%m-%d')
            s = str(val)[:10] if val else None
            return s if s and validate_date(s) else None

        def _safe_float(val, default=0):
            if val is None:
                return default
            try:
                return float(val)
            except (ValueError, TypeError):
                return default

        def _safe_pct(val, default=1.0):
            f = _safe_float(val, default)
            return max(0.0, min(f, 1.0))

        def _safe_str(val, max_len=500):
            if val is None:
                return None
            s = str(val)[:max_len]
            return s

        with get_db() as conn:
            # Positions
            ws = wb['Positions']
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if i >= MAX_IMPORT_ROWS:
                    break
                if len(row) < 7:
                    continue
                date_val, owner, category = row[0], row[1], row[2]
                envelope, establishment   = row[3], row[4]
                value, debt               = row[5], row[6]
                notes                     = row[8] if len(row) > 8 else None
                entity                    = row[9] if len(row) > 9 else None
                ownership_pct_raw = row[10] if len(row) > 10 else None
                debt_pct_raw      = row[11] if len(row) > 11 else None

                date_str = _parse_date(date_val)
                if not date_str or not owner:
                    skipped += 1
                    continue
                if value is None and debt is None and envelope is None and not entity:
                    continue

                owner    = _safe_str(owner, 100)
                category = _safe_str(category, 100) or ''
                envelope = _safe_str(envelope, 100)
                establishment = _safe_str(establishment, 200)
                notes    = _safe_str(notes, MAX_NOTE_LENGTH)
                entity   = _safe_str(entity, 200)

                if entity:
                    value = 0
                    debt  = 0
                    ownership_pct = _safe_pct(ownership_pct_raw, 1.0)
                    debt_pct = _safe_pct(debt_pct_raw, ownership_pct)
                else:
                    value = _safe_float(value, 0)
                    debt  = _safe_float(debt, 0)
                    ownership_pct = _safe_pct(ownership_pct_raw, 1.0)
                    debt_pct      = _safe_pct(debt_pct_raw, 1.0)

                existing = conn.execute(
                    '''SELECT id FROM positions
                       WHERE date=? AND owner=? AND category=?
                         AND COALESCE(envelope,'')=? AND COALESCE(entity,'')=?''',
                    (date_str, owner, category, envelope or '', entity or '')
                ).fetchone()
                if existing:
                    continue

                conn.execute(
                    '''INSERT INTO positions
                       (date, owner, category, envelope, establishment,
                        value, debt, notes, entity, ownership_pct, debt_pct)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                    (date_str, owner, category,
                     envelope, establishment,
                     value, debt,
                     notes, entity,
                     ownership_pct, debt_pct)
                )
                imported += 1

            # Flux
            if 'Flux' in wb.sheetnames:
                wf = wb['Flux']
                flux_imported = 0
                for i, row in enumerate(wf.iter_rows(min_row=2, values_only=True)):
                    if i >= MAX_IMPORT_ROWS:
                        break
                    if len(row) < 5:
                        continue
                    date_val, owner, envelope, ftype, amount = row[0], row[1], row[2], row[3], row[4]
                    notes = row[5] if len(row) > 5 else None
                    date_str = _parse_date(date_val)
                    if not date_str or not owner or amount is None:
                        continue
                    conn.execute(
                        'INSERT INTO flux (date, owner, envelope, type, amount, notes) VALUES (?,?,?,?,?,?)',
                        (date_str, _safe_str(owner, 100), _safe_str(envelope, 100),
                         _safe_str(ftype, 50), _safe_float(amount, 0), _safe_str(notes, MAX_NOTE_LENGTH))
                    )
                    flux_imported += 1

            # Entités
            entities_imported = 0
            if 'Entites' in wb.sheetnames:
                we = wb['Entites']
                for i, row in enumerate(we.iter_rows(min_row=2, values_only=True)):
                    if i >= 1000:
                        break
                    if len(row) < 2:
                        continue
                    name = row[0]
                    if not name:
                        continue
                    name           = _safe_str(name, 200)
                    etype          = _safe_str(row[1], 50) if len(row) > 1 else None
                    valuation_mode = _safe_str(row[2], 50) if len(row) > 2 else None
                    gross_assets   = _safe_float(row[3] if len(row) > 3 else 0)
                    debt           = _safe_float(row[4] if len(row) > 4 else 0)
                    comment        = _safe_str(row[6], MAX_NOTE_LENGTH) if len(row) > 6 else None
                    existing = conn.execute(
                        'SELECT id FROM entities WHERE name=?', (name,)
                    ).fetchone()
                    if existing:
                        conn.execute(
                            '''UPDATE entities SET type=?, valuation_mode=?,
                               gross_assets=?, debt=?, comment=? WHERE name=?''',
                            (etype, valuation_mode, gross_assets, debt, comment, name)
                        )
                    else:
                        conn.execute(
                            '''INSERT INTO entities (name, type, valuation_mode, gross_assets, debt, comment)
                               VALUES (?,?,?,?,?,?)''',
                            (name, etype, valuation_mode, gross_assets, debt, comment)
                        )
                    today = datetime.now().strftime('%Y-%m-%d')
                    conn.execute(
                        '''INSERT OR REPLACE INTO entity_snapshots (entity_name, date, gross_assets, debt)
                           VALUES (?,?,?,?)''',
                        (name, today, gross_assets, debt)
                    )
                    entities_imported += 1

            # Securities (optionnel, avant Holdings pour que les FK existent)
            securities_imported = 0
            if 'Securities' in wb.sheetnames:
                ws_sec = wb['Securities']
                for i, row in enumerate(ws_sec.iter_rows(min_row=2, values_only=True)):
                    if i >= MAX_IMPORT_ROWS:
                        break
                    if not row or not row[0]:
                        continue
                    raw_isin = str(row[0]).strip().upper()
                    if not validate_isin(raw_isin):
                        continue
                    name         = _safe_str(row[1] if len(row) > 1 else None, 200)
                    ticker       = _safe_str(row[2] if len(row) > 2 else None, 50)
                    currency     = _safe_str(row[3] if len(row) > 3 else 'EUR', 10) or 'EUR'
                    asset_class  = _safe_str(row[4] if len(row) > 4 else None, 50)
                    is_priceable = row[5] if len(row) > 5 else 1
                    is_priceable = 0 if str(is_priceable).lower() in ('0', 'false', 'non', 'no', '') else 1
                    if raw_isin.startswith(('FONDS_EUROS_', 'CUSTOM_')):
                        is_priceable = 0
                    existing = conn.execute(
                        'SELECT isin FROM securities WHERE isin=?', (raw_isin,)
                    ).fetchone()
                    if existing:
                        conn.execute(
                            '''UPDATE securities SET name=?, ticker=?, currency=?,
                               asset_class=?, is_priceable=?, updated_at=CURRENT_TIMESTAMP
                               WHERE isin=?''',
                            (name, ticker, currency, asset_class, is_priceable, raw_isin)
                        )
                    else:
                        conn.execute(
                            '''INSERT INTO securities
                               (isin, name, ticker, currency, asset_class, is_priceable, data_source)
                               VALUES (?,?,?,?,?,?,'xlsx')''',
                            (raw_isin, name, ticker, currency, asset_class, is_priceable)
                        )
                    securities_imported += 1

            # Holdings : rattachement par clef (date, owner, category, envelope, entity)
            holdings_imported = 0
            if 'Holdings' in wb.sheetnames:
                ws_h = wb['Holdings']
                for i, row in enumerate(ws_h.iter_rows(min_row=2, values_only=True)):
                    if i >= MAX_IMPORT_ROWS:
                        break
                    if not row or len(row) < 7:
                        continue
                    pos_date = _parse_date(row[0])
                    pos_owner = _safe_str(row[1], 100)
                    pos_category = _safe_str(row[2], 100) or ''
                    pos_envelope = _safe_str(row[3], 100) or ''
                    pos_entity = _safe_str(row[4], 200) or ''
                    raw_isin = str(row[5]).strip().upper() if row[5] else ''
                    quantity = _safe_float(row[6], 0)
                    cost_basis = _safe_float(row[7], 0) if len(row) > 7 and row[7] is not None else None
                    market_value = _safe_float(row[8], 0) if len(row) > 8 and row[8] is not None else None
                    as_of_date = _parse_date(row[9]) if len(row) > 9 else None

                    if not pos_date or not pos_owner or not raw_isin or quantity <= 0:
                        continue
                    if not validate_isin(raw_isin):
                        continue

                    pos = conn.execute(
                        '''SELECT id FROM positions
                           WHERE date=? AND owner=? AND category=?
                             AND COALESCE(envelope,'')=? AND COALESCE(entity,'')=?''',
                        (pos_date, pos_owner, pos_category, pos_envelope, pos_entity)
                    ).fetchone()
                    if not pos:
                        continue  # Position parente introuvable → skip

                    # Upsert auto de la security si absente
                    existing = conn.execute(
                        'SELECT isin FROM securities WHERE isin=?', (raw_isin,)
                    ).fetchone()
                    if not existing:
                        is_priceable = 0 if raw_isin.startswith(('FONDS_EUROS_', 'CUSTOM_')) else 1
                        conn.execute(
                            '''INSERT INTO securities
                               (isin, currency, is_priceable, data_source)
                               VALUES (?,'EUR',?,'xlsx-auto')''',
                            (raw_isin, is_priceable)
                        )
                    conn.execute(
                        '''INSERT INTO holdings
                           (position_id, isin, quantity, cost_basis, market_value, as_of_date)
                           VALUES (?,?,?,?,?,?)''',
                        (pos['id'], raw_isin, quantity, cost_basis, market_value, as_of_date)
                    )
                    holdings_imported += 1

        logger.info('Import XLSX: %d positions, %d entités, %d securities, %d holdings, %d skipped',
                    imported, entities_imported, securities_imported, holdings_imported, skipped)
        return jsonify({
            'imported': imported,
            'entities': entities_imported,
            'securities': securities_imported,
            'holdings': holdings_imported,
            'skipped': skipped,
        })

    except Exception as e:
        logger.error('Import XLSX failed: %s', e, exc_info=True)
        return jsonify({'error': "Échec de l'import — vérifiez le format du fichier."}), 400


@import_export_bp.route('/api/import-json', methods=['POST'])
@login_required
@csrf_protect
def import_json():
    data = request.json
    if not data:
        return jsonify({'error': 'Corps JSON manquant'}), 400
    if not isinstance(data, dict):
        return jsonify({'error': 'Objet JSON attendu'}), 400

    imported = {'positions': 0, 'flux': 0, 'entities': 0, 'entity_snapshots': 0,
                'securities': 0, 'holdings': 0, 'holdings_snapshots': 0, 'skipped': 0}

    def _clamp_pct(v, default=1.0):
        try:
            f = float(v) if v is not None else default
        except (ValueError, TypeError):
            return default
        return max(0.0, min(f, 1.0))

    def _safe_num(v, default=0):
        try:
            return float(v) if v is not None else default
        except (ValueError, TypeError):
            return default

    def _trunc(v, max_len):
        if v is None:
            return None
        return str(v)[:max_len]

    with get_db() as conn:
        for e in data.get('entities', [])[:1000]:
            name = _trunc(e.get('name'), 200)
            if not name:
                continue
            existing = conn.execute('SELECT id FROM entities WHERE name=?', (name,)).fetchone()
            if existing:
                conn.execute(
                    'UPDATE entities SET type=?, valuation_mode=?, gross_assets=?, debt=?, comment=? WHERE name=?',
                    (_trunc(e.get('type'), 50), _trunc(e.get('valuation_mode'), 50),
                     _safe_num(e.get('gross_assets')), _safe_num(e.get('debt')),
                     _trunc(e.get('comment'), MAX_NOTE_LENGTH), name)
                )
            else:
                conn.execute(
                    'INSERT INTO entities (name, type, valuation_mode, gross_assets, debt, comment) VALUES (?,?,?,?,?,?)',
                    (name, _trunc(e.get('type'), 50), _trunc(e.get('valuation_mode'), 50),
                     _safe_num(e.get('gross_assets')), _safe_num(e.get('debt')),
                     _trunc(e.get('comment'), MAX_NOTE_LENGTH))
                )
            imported['entities'] += 1

        for s in data.get('entity_snapshots', [])[:MAX_IMPORT_ROWS]:
            ename = _trunc(s.get('entity_name'), 200)
            sdate = s.get('date')
            if not ename or not validate_date(sdate):
                imported['skipped'] += 1
                continue
            conn.execute(
                'INSERT OR REPLACE INTO entity_snapshots (entity_name, date, gross_assets, debt) VALUES (?,?,?,?)',
                (ename, sdate, _safe_num(s.get('gross_assets')), _safe_num(s.get('debt')))
            )
            imported['entity_snapshots'] += 1

        for p in data.get('positions', [])[:MAX_IMPORT_ROWS]:
            pdate = p.get('date')
            owner = _trunc(p.get('owner'), 100)
            if not validate_date(pdate) or not owner:
                imported['skipped'] += 1
                continue
            existing = conn.execute(
                '''SELECT id FROM positions WHERE date=? AND owner=? AND category=?
                   AND COALESCE(envelope,'')=? AND COALESCE(entity,'')=?''',
                (pdate, owner, _trunc(p.get('category'), 100) or '', p.get('envelope') or '', p.get('entity') or '')
            ).fetchone()
            if existing:
                continue
            conn.execute(
                '''INSERT INTO positions (date, owner, category, envelope, establishment,
                   value, debt, notes, entity, ownership_pct, debt_pct) VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (pdate, owner, _trunc(p.get('category'), 100) or '',
                 _trunc(p.get('envelope'), 100), _trunc(p.get('establishment'), 200),
                 _safe_num(p.get('value')), _safe_num(p.get('debt')),
                 _trunc(p.get('notes'), MAX_NOTE_LENGTH), _trunc(p.get('entity'), 200),
                 _clamp_pct(p.get('ownership_pct')), _clamp_pct(p.get('debt_pct')))
            )
            imported['positions'] += 1

        for f in data.get('flux', [])[:MAX_IMPORT_ROWS]:
            fdate = f.get('date')
            fowner = _trunc(f.get('owner'), 100)
            if not validate_date(fdate) or not fowner or f.get('amount') is None:
                imported['skipped'] += 1
                continue
            conn.execute(
                'INSERT INTO flux (date, owner, envelope, type, amount, notes, category) VALUES (?,?,?,?,?,?,?)',
                (fdate, fowner, _trunc(f.get('envelope'), 100), _trunc(f.get('type'), 50),
                 _safe_num(f.get('amount')), _trunc(f.get('notes'), MAX_NOTE_LENGTH),
                 _trunc(f.get('category'), 100))
            )
            imported['flux'] += 1

        for date, notes in data.get('snapshot_notes', {}).items():
            if validate_date(date) and notes:
                conn.execute(
                    'INSERT OR REPLACE INTO snapshot_notes (date, notes) VALUES (?, ?)',
                    (date, str(notes)[:MAX_NOTE_LENGTH])
                )

        # Securities (obligatoire avant holdings)
        for sec in data.get('securities', [])[:MAX_IMPORT_ROWS]:
            isin = str(sec.get('isin') or '').strip().upper()
            if not validate_isin(isin):
                imported['skipped'] += 1
                continue
            is_priceable = sec.get('is_priceable')
            if is_priceable is None:
                is_priceable = 0 if isin.startswith(('FONDS_EUROS_', 'CUSTOM_')) else 1
            else:
                is_priceable = 0 if not is_priceable else 1
            existing = conn.execute('SELECT isin FROM securities WHERE isin=?', (isin,)).fetchone()
            if existing:
                conn.execute(
                    '''UPDATE securities SET name=?, ticker=?, currency=?, asset_class=?,
                       is_priceable=?, last_price=?, last_price_date=?, updated_at=CURRENT_TIMESTAMP
                       WHERE isin=?''',
                    (_trunc(sec.get('name'), 200), _trunc(sec.get('ticker'), 50),
                     _trunc(sec.get('currency'), 10) or 'EUR',
                     _trunc(sec.get('asset_class'), 50), is_priceable,
                     _safe_num(sec.get('last_price')) if sec.get('last_price') is not None else None,
                     sec.get('last_price_date') if validate_date(sec.get('last_price_date')) else None,
                     isin)
                )
            else:
                conn.execute(
                    '''INSERT INTO securities
                       (isin, name, ticker, currency, asset_class, is_priceable,
                        last_price, last_price_date, data_source)
                       VALUES (?,?,?,?,?,?,?,?,?)''',
                    (isin, _trunc(sec.get('name'), 200), _trunc(sec.get('ticker'), 50),
                     _trunc(sec.get('currency'), 10) or 'EUR',
                     _trunc(sec.get('asset_class'), 50), is_priceable,
                     _safe_num(sec.get('last_price')) if sec.get('last_price') is not None else None,
                     sec.get('last_price_date') if validate_date(sec.get('last_price_date')) else None,
                     _trunc(sec.get('data_source'), 50) or 'json')
                )
            imported['securities'] += 1

        # Holdings : matchés sur la clef métier de position
        for h in data.get('holdings', [])[:MAX_IMPORT_ROWS]:
            isin = str(h.get('isin') or '').strip().upper()
            if not validate_isin(isin):
                imported['skipped'] += 1
                continue
            pos_date = h.get('pos_date') or h.get('date')
            if not validate_date(pos_date):
                imported['skipped'] += 1
                continue
            try:
                qty = float(h.get('quantity') or 0)
            except (ValueError, TypeError):
                imported['skipped'] += 1
                continue
            if qty <= 0:
                imported['skipped'] += 1
                continue
            row = conn.execute(
                '''SELECT id FROM positions WHERE date=? AND owner=? AND category=?
                   AND COALESCE(envelope,'')=? AND COALESCE(entity,'')=?''',
                (pos_date, _trunc(h.get('pos_owner'), 100),
                 _trunc(h.get('pos_category'), 100) or '',
                 _trunc(h.get('pos_envelope'), 100) or '',
                 _trunc(h.get('pos_entity'), 200) or '')
            ).fetchone()
            if not row:
                imported['skipped'] += 1
                continue
            # Auto-upsert security si inconnue
            if not conn.execute('SELECT 1 FROM securities WHERE isin=?', (isin,)).fetchone():
                is_priceable = 0 if isin.startswith(('FONDS_EUROS_', 'CUSTOM_')) else 1
                conn.execute(
                    '''INSERT INTO securities (isin, currency, is_priceable, data_source)
                       VALUES (?,'EUR',?, 'json-auto')''',
                    (isin, is_priceable)
                )
            conn.execute(
                '''INSERT INTO holdings
                   (position_id, isin, quantity, cost_basis, market_value, as_of_date)
                   VALUES (?,?,?,?,?,?)''',
                (row['id'], isin, qty,
                 _safe_num(h.get('cost_basis')) if h.get('cost_basis') is not None else None,
                 _safe_num(h.get('market_value')) if h.get('market_value') is not None else None,
                 h.get('as_of_date') if validate_date(h.get('as_of_date')) else None)
            )
            imported['holdings'] += 1

        # Holdings snapshots (reconstruction de l'historique détaillé)
        for s in data.get('holdings_snapshots', [])[:MAX_IMPORT_ROWS]:
            snap_date = s.get('snapshot_date')
            isin = str(s.get('isin') or '').strip().upper()
            if not validate_date(snap_date) or not validate_isin(isin):
                imported['skipped'] += 1
                continue
            position_id = s.get('position_id')
            # Skip si la position cible n'existe plus (import cross-DB cassé)
            if position_id is not None:
                exists = conn.execute(
                    'SELECT 1 FROM positions WHERE id=?', (position_id,)
                ).fetchone()
                if not exists:
                    imported['skipped'] += 1
                    continue
            conn.execute(
                '''INSERT INTO holdings_snapshots
                   (snapshot_date, position_id, isin, quantity, cost_basis, price, market_value)
                   VALUES (?,?,?,?,?,?,?)''',
                (snap_date, position_id, isin,
                 _safe_num(s.get('quantity')),
                 _safe_num(s.get('cost_basis')) if s.get('cost_basis') is not None else None,
                 _safe_num(s.get('price')) if s.get('price') is not None else None,
                 _safe_num(s.get('market_value')) if s.get('market_value') is not None else None)
            )
            imported['holdings_snapshots'] += 1

    logger.info('Import JSON: %s', imported)
    return jsonify(imported)


@import_export_bp.route('/api/export')
@login_required
def export_data():
    with get_db() as conn:
        positions = [dict(r) for r in conn.execute(
            'SELECT * FROM positions ORDER BY date, owner'
        ).fetchall()]
        flux = [dict(r) for r in conn.execute(
            'SELECT * FROM flux ORDER BY date'
        ).fetchall()]
        entities = [dict(r) for r in conn.execute(
            'SELECT * FROM entities ORDER BY name'
        ).fetchall()]
        entity_snapshots = [dict(r) for r in conn.execute(
            'SELECT * FROM entity_snapshots ORDER BY entity_name, date'
        ).fetchall()]
        snapshot_notes = {r['date']: r['notes'] for r in conn.execute(
            'SELECT date, notes FROM snapshot_notes ORDER BY date'
        ).fetchall()}

        securities, holdings, holdings_snapshots = [], [], []
        try:
            securities = [dict(r) for r in conn.execute(
                'SELECT * FROM securities ORDER BY isin'
            ).fetchall()]
            # Pour les holdings, on exporte la clef métier de la position plutôt
            # que l'id (qui ne survit pas un reset/reimport).
            holdings = [dict(r) for r in conn.execute(
                '''SELECT h.isin, h.quantity, h.cost_basis, h.market_value, h.as_of_date,
                          p.date AS pos_date, p.owner AS pos_owner, p.category AS pos_category,
                          COALESCE(p.envelope,'')  AS pos_envelope,
                          COALESCE(p.entity,'')    AS pos_entity
                   FROM holdings h
                   JOIN positions p ON p.id = h.position_id
                   ORDER BY p.date, p.owner, h.id'''
            ).fetchall()]
            holdings_snapshots = [dict(r) for r in conn.execute(
                'SELECT * FROM holdings_snapshots ORDER BY snapshot_date, id'
            ).fetchall()]
        except Exception:
            pass  # tables holdings non migrées

    return jsonify({
        'positions': positions,
        'flux': flux,
        'entities': entities,
        'entity_snapshots': entity_snapshots,
        'snapshot_notes': snapshot_notes,
        'securities': securities,
        'holdings': holdings,
        'holdings_snapshots': holdings_snapshots,
    })


@import_export_bp.route('/api/reset', methods=['POST'])
@login_required
@csrf_protect
def reset_db():
    tables = [
        'positions', 'flux', 'entities', 'entity_snapshots', 'snapshot_notes',
        'holdings', 'holdings_snapshots', 'price_history', 'securities',
    ]
    deleted = {}
    with get_db() as conn:
        for t in tables:
            try:
                row = conn.execute(f'SELECT COUNT(*) AS c FROM {t}').fetchone()
                deleted[t] = int(row['c']) if row else 0
                conn.execute(f'DELETE FROM {t}')
            except Exception:
                deleted[t] = 0  # Table peut ne pas exister si migration non appliquée
    total = sum(deleted.values())
    logger.warning('Database reset — %d rows across %d tables deleted',
                   total, len([v for v in deleted.values() if v]))
    return jsonify({'ok': True, 'total': total, 'deleted': deleted})
