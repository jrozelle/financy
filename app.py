from flask import Flask, jsonify, request, render_template, session, redirect, url_for
from functools import wraps
import sqlite3
import json
import os
import re
from datetime import datetime
from contextlib import contextmanager
from io import BytesIO

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), 'patrimoine.db')
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(32))

# ─── Authentification ─────────────────────────────────────────────────────────

AUTH_PASSWORD = os.environ.get('FINANCY_PASSWORD')  # None = pas d'auth

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if AUTH_PASSWORD and not session.get('authenticated'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Non authentifié'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if not AUTH_PASSWORD:
        return redirect(url_for('index'))
    if request.method == 'POST':
        if request.form.get('password') == AUTH_PASSWORD:
            session['authenticated'] = True
            return redirect(url_for('index'))
        return render_template('login.html', error='Mot de passe incorrect.')
    return render_template('login.html', error=None)

@app.route('/logout')
def logout():
    session.pop('authenticated', None)
    return redirect(url_for('login_page'))


# ─── Validation ───────────────────────────────────────────────────────────────

def validate_date(s):
    """Vérifie que la chaîne est une date ISO valide (YYYY-MM-DD)."""
    if not s or not isinstance(s, str):
        return False
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return False
    try:
        datetime.strptime(s, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def validate_number(v, allow_negative=False):
    """Vérifie que v est un nombre valide."""
    if v is None:
        return True
    try:
        n = float(v)
        if not allow_negative and n < 0:
            return False
        return True
    except (ValueError, TypeError):
        return False

def validate_string(s, max_length=500):
    """Vérifie que s est une chaîne non vide et raisonnable."""
    if s is None:
        return True
    return isinstance(s, str) and len(s) <= max_length

def validate_pct(v):
    """Vérifie que v est un pourcentage entre 0 et 1."""
    if v is None:
        return True
    try:
        n = float(v)
        return 0 <= n <= 1.0001  # petite marge pour les arrondis
    except (ValueError, TypeError):
        return False

# ─── Référentiels ────────────────────────────────────────────────────────────

OWNERS = ['Julien', 'Perrine', 'Adriel', 'Aloïs']

CATEGORIES = [
    'Cash & dépôts', 'Monétaire', 'Obligations', 'Actions',
    'Immobilier', 'SCPI', 'Fond Euro', 'Produits Structurés',
    'Crypto', 'Objets de valeur', 'Autre'
]

ENVELOPE_META = {
    'Compte courant':  {'liquidity': 'J0–J1',  'friction': 'Aucune'},
    'Livret A':        {'liquidity': 'J2–J7',  'friction': 'Fiscale'},
    'LDDS':            {'liquidity': 'J0–J1',  'friction': 'Aucune'},
    'Livret Bourso+':  {'liquidity': 'J0–J1',  'friction': 'Aucune'},
    'PEL/CEL':         {'liquidity': 'J8–J30', 'friction': 'Frais'},
    'PEA':             {'liquidity': 'J2–J7',  'friction': 'Fiscale'},
    'CTO':             {'liquidity': 'J2–J7',  'friction': 'Fiscale'},
    'Assurance-vie':   {'liquidity': 'J8–J30', 'friction': 'Mixte'},
    'PER':             {'liquidity': 'Bloqué', 'friction': 'Fiscale'},
    'Crypto':          {'liquidity': 'J0–J1',  'friction': 'Décote probable'},
    'Immobilier':      {'liquidity': '30J+',   'friction': 'Mixte'},
    'SCI':             {'liquidity': '30J+',   'friction': 'Mixte'},
    'Dette':           {'liquidity': 'Bloqué', 'friction': 'Aucune'},
    'Autre':           {'liquidity': '30J+',   'friction': 'Mixte'},
}

CATEGORY_MOBILIZABLE = {
    'Cash & dépôts':       1.0,
    'Monétaire':           0.95,
    'Obligations':         0.95,
    'Actions':             0.9,
    'Immobilier':          0.0,
    'SCPI':                0.0,
    'Fond Euro':           0.95,
    'Produits Structurés': 0.0,
    'Crypto':              0.9,
    'Objets de valeur':    0.0,
    'Autre':               0.8,
}

FLUX_TYPES = ['Versement', 'Retrait', 'Dividende/Intérêt', 'Frais', 'Autre']
LIQUIDITY_ORDER = ['J0–J1', 'J2–J7', 'J8–J30', '30J+', 'Bloqué']
ENTITY_TYPES = ['SCI', 'Indivision', 'Holding', 'Autre']
VALUATION_MODES = ['Valeur de marché', "Prix d'acquisition", 'Valeur fiscale', 'Autre']

# ─── Référentiel dynamique ───────────────────────────────────────────────────

DEFAULT_REFERENTIAL = {
    'owners':               OWNERS,
    'categories':           CATEGORIES,
    'category_mobilizable': CATEGORY_MOBILIZABLE,
    'envelope_meta':        ENVELOPE_META,
    'entity_types':         ENTITY_TYPES,
    'valuation_modes':      VALUATION_MODES,
    'flux_types':           FLUX_TYPES,
    'liquidity_order':      LIQUIDITY_ORDER,
}

def load_referential(conn):
    """Charge le référentiel depuis la DB ; fallback sur les defaults."""
    row = conn.execute("SELECT value FROM config WHERE key='referential'").fetchone()
    if row:
        try:
            stored = json.loads(row['value'])
            ref = dict(DEFAULT_REFERENTIAL)
            ref.update(stored)
            # liquidity_order n'est pas éditable via l'UI → toujours la valeur par défaut
            ref['liquidity_order'] = LIQUIDITY_ORDER
            return ref
        except Exception:
            pass
    return dict(DEFAULT_REFERENTIAL)

# ─── Base de données ─────────────────────────────────────────────────────────

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db():
    with get_db() as conn:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS positions (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                date            TEXT    NOT NULL,
                owner           TEXT    NOT NULL,
                category        TEXT    NOT NULL,
                envelope        TEXT,
                establishment   TEXT,
                value           REAL    DEFAULT 0,
                debt            REAL    DEFAULT 0,
                notes           TEXT,
                entity          TEXT,
                ownership_pct   REAL    DEFAULT 1.0,
                debt_pct        REAL    DEFAULT 1.0,
                created_at      TEXT    DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS entities (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                name             TEXT    NOT NULL UNIQUE,
                type             TEXT,
                valuation_mode   TEXT,
                gross_assets     REAL    DEFAULT 0,
                debt             REAL    DEFAULT 0,
                comment          TEXT,
                created_at       TEXT    DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS flux (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                date        TEXT    NOT NULL,
                owner       TEXT    NOT NULL,
                envelope    TEXT,
                type        TEXT,
                amount      REAL    NOT NULL,
                notes       TEXT,
                created_at  TEXT    DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX IF NOT EXISTS idx_positions_date ON positions(date);
            CREATE INDEX IF NOT EXISTS idx_flux_date      ON flux(date);
            CREATE TABLE IF NOT EXISTS config (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS entity_snapshots (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                entity_name  TEXT NOT NULL,
                date         TEXT NOT NULL,
                gross_assets REAL DEFAULT 0,
                debt         REAL DEFAULT 0,
                created_at   TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(entity_name, date)
            );
            CREATE INDEX IF NOT EXISTS idx_entity_snap ON entity_snapshots(entity_name, date);
        ''')
        # Migration : colonne ajoutée après la création initiale
        try:
            conn.execute('ALTER TABLE positions ADD COLUMN mobilizable_pct_override REAL DEFAULT NULL')
        except Exception:
            pass  # colonne déjà présente

# ─── Calculs ─────────────────────────────────────────────────────────────────

def compute_position(pos, entity_map=None, ref=None):
    """
    entity_map : dict {entity_name: {gross_assets, debt}} optionnel.
    ref        : référentiel dynamique (load_referential). Défaut : DEFAULT_REFERENTIAL.
    """
    if ref is None:
        ref = DEFAULT_REFERENTIAL
    ownership_pct = pos.get('ownership_pct') if pos.get('ownership_pct') is not None else 1.0
    debt_pct      = pos.get('debt_pct')      if pos.get('debt_pct')      is not None else 1.0
    category      = pos.get('category', '')
    envelope      = pos.get('envelope', '') or ''
    entity        = pos.get('entity')

    if entity and entity_map and entity in entity_map:
        value = entity_map[entity]['gross_assets'] or 0
        debt  = entity_map[entity]['debt'] or 0
    else:
        value = pos.get('value') or 0
        debt  = pos.get('debt') or 0

    gross_attributed = value * ownership_pct
    debt_attributed  = debt * debt_pct
    net_attributed   = gross_attributed - debt_attributed

    env_meta         = ref.get('envelope_meta', ENVELOPE_META)
    cat_mob          = ref.get('category_mobilizable', CATEGORY_MOBILIZABLE)
    env              = env_meta.get(envelope, {'liquidity': '30J+', 'friction': 'Mixte'})
    override         = pos.get('mobilizable_pct_override')
    mobilizable_pct  = override if override is not None else cat_mob.get(category, 0.8)
    mobilizable_val  = net_attributed * mobilizable_pct if net_attributed > 0 else 0

    return {
        **pos,
        'net_value':         value - debt,
        'gross_attributed':  gross_attributed,
        'debt_attributed':   debt_attributed,
        'net_attributed':    net_attributed,
        'liquidity':         env['liquidity'],
        'friction':          env['friction'],
        'mobilizable_pct':   mobilizable_pct,
        'mobilizable_value': mobilizable_val,
    }

def get_entity_map(conn, date=None):
    """
    Retourne {name: {gross_assets, debt}} pour toutes les entités.
    Si date est fourni, utilise la valorisation la plus récente <= date
    (versionnement historique). Fallback sur la valeur courante de l'entité.
    """
    if date:
        rows = conn.execute('''
            SELECT e.name,
                   COALESCE(s.gross_assets, e.gross_assets) AS gross_assets,
                   COALESCE(s.debt,         e.debt)         AS debt
            FROM entities e
            LEFT JOIN entity_snapshots s
              ON s.entity_name = e.name
             AND s.date = (
                 SELECT MAX(date) FROM entity_snapshots es2
                 WHERE es2.entity_name = e.name AND es2.date <= ?
             )
        ''', (date,)).fetchall()
    else:
        rows = conn.execute('SELECT name, gross_assets, debt FROM entities').fetchall()
    return {r['name']: {'gross_assets': r['gross_assets'] or 0, 'debt': r['debt'] or 0}
            for r in rows}

# ─── Routes API ──────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    return render_template('index.html', has_auth=bool(AUTH_PASSWORD))


@app.route('/api/config')
@login_required
def get_config():
    with get_db() as conn:
        entity_names = [r['name'] for r in
                        conn.execute('SELECT name FROM entities ORDER BY name').fetchall()]
        ref = load_referential(conn)
    return jsonify({
        'owners':               ref['owners'],
        'categories':           ref['categories'],
        'envelopes':            list(ref['envelope_meta'].keys()),
        'flux_types':           ref['flux_types'],
        'liquidity_order':      ref['liquidity_order'],
        'category_mobilizable': ref['category_mobilizable'],
        'envelope_meta':        ref['envelope_meta'],
        'entity_types':         ref['entity_types'],
        'valuation_modes':      ref['valuation_modes'],
        'entity_names':         entity_names,
    })


@app.route('/api/dates')
@login_required
def get_dates():
    with get_db() as conn:
        rows = conn.execute(
            'SELECT DISTINCT date FROM positions ORDER BY date DESC'
        ).fetchall()
    return jsonify([r['date'] for r in rows])


# — Positions —

@app.route('/api/positions', methods=['GET'])
@login_required
def get_positions():
    date = request.args.get('date')
    with get_db() as conn:
        if date:
            rows = conn.execute(
                'SELECT * FROM positions WHERE date=? ORDER BY owner, category', (date,)
            ).fetchall()
        else:
            rows = conn.execute(
                'SELECT * FROM positions ORDER BY date DESC, owner, category'
            ).fetchall()
        entity_map = get_entity_map(conn, date)
        ref        = load_referential(conn)
    return jsonify([compute_position(dict(r), entity_map, ref) for r in rows])


@app.route('/api/positions', methods=['POST'])
@login_required
def add_position():
    d = request.json
    if not d or not validate_date(d.get('date')):
        return jsonify({'error': 'Date invalide (format AAAA-MM-JJ attendu)'}), 400
    if not validate_string(d.get('owner'), 100) or not d.get('owner'):
        return jsonify({'error': 'Propriétaire requis'}), 400
    if not validate_string(d.get('category'), 100) or not d.get('category'):
        return jsonify({'error': 'Catégorie requise'}), 400
    if not validate_number(d.get('value')) or not validate_number(d.get('debt')):
        return jsonify({'error': 'Valeur / dette invalide'}), 400
    if not validate_pct(d.get('ownership_pct')) or not validate_pct(d.get('debt_pct')):
        return jsonify({'error': '% propriété ou dette invalide (0-100)'}), 400
    with get_db() as conn:
        entity = d.get('entity')
        stored_value = 0 if entity else d.get('value', 0)
        stored_debt  = 0 if entity else d.get('debt', 0)
        mob_override = d.get('mobilizable_pct_override')
        if mob_override is not None:
            mob_override = float(mob_override)
        cur = conn.execute(
            '''INSERT INTO positions
               (date, owner, category, envelope, establishment, value, debt,
                notes, entity, ownership_pct, debt_pct, mobilizable_pct_override)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (d['date'], d['owner'], d['category'],
             d.get('envelope'), d.get('establishment'),
             stored_value, stored_debt,
             d.get('notes'), entity,
             d.get('ownership_pct', 1.0), d.get('debt_pct', 1.0), mob_override)
        )
        row        = conn.execute('SELECT * FROM positions WHERE id=?', (cur.lastrowid,)).fetchone()
        entity_map = get_entity_map(conn)
        ref        = load_referential(conn)
    return jsonify(compute_position(dict(row), entity_map, ref)), 201


@app.route('/api/positions/<int:pid>', methods=['PUT'])
@login_required
def update_position(pid):
    d = request.json
    if not d or not validate_date(d.get('date')):
        return jsonify({'error': 'Date invalide'}), 400
    if not validate_number(d.get('value')) or not validate_number(d.get('debt')):
        return jsonify({'error': 'Valeur / dette invalide'}), 400
    if not validate_pct(d.get('ownership_pct')) or not validate_pct(d.get('debt_pct')):
        return jsonify({'error': '% invalide'}), 400
    with get_db() as conn:
        entity = d.get('entity')
        stored_value = 0 if entity else d.get('value', 0)
        stored_debt  = 0 if entity else d.get('debt', 0)
        mob_override = d.get('mobilizable_pct_override')
        if mob_override is not None:
            mob_override = float(mob_override)
        conn.execute(
            '''UPDATE positions SET
               date=?, owner=?, category=?, envelope=?, establishment=?,
               value=?, debt=?, notes=?, entity=?, ownership_pct=?, debt_pct=?,
               mobilizable_pct_override=?
               WHERE id=?''',
            (d['date'], d['owner'], d['category'],
             d.get('envelope'), d.get('establishment'),
             stored_value, stored_debt,
             d.get('notes'), entity,
             d.get('ownership_pct', 1.0), d.get('debt_pct', 1.0), mob_override, pid)
        )
        row        = conn.execute('SELECT * FROM positions WHERE id=?', (pid,)).fetchone()
        entity_map = get_entity_map(conn)
        ref        = load_referential(conn)
    return jsonify(compute_position(dict(row), entity_map, ref))


@app.route('/api/positions/<int:pid>', methods=['DELETE'])
@login_required
def delete_position(pid):
    with get_db() as conn:
        conn.execute('DELETE FROM positions WHERE id=?', (pid,))
    return '', 204


# — Flux —

@app.route('/api/flux', methods=['GET'])
@login_required
def get_flux():
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')
    with get_db() as conn:
        if date_from and date_to:
            rows = conn.execute(
                'SELECT * FROM flux WHERE date >= ? AND date <= ? ORDER BY date DESC',
                (date_from, date_to)
            ).fetchall()
        else:
            rows = conn.execute('SELECT * FROM flux ORDER BY date DESC').fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/flux', methods=['POST'])
@login_required
def add_flux():
    d = request.json
    if not d or not validate_date(d.get('date')):
        return jsonify({'error': 'Date invalide'}), 400
    if not d.get('owner'):
        return jsonify({'error': 'Propriétaire requis'}), 400
    if not validate_number(d.get('amount'), allow_negative=True) or d.get('amount') is None:
        return jsonify({'error': 'Montant invalide'}), 400
    with get_db() as conn:
        cur = conn.execute(
            'INSERT INTO flux (date, owner, envelope, type, amount, notes) VALUES (?,?,?,?,?,?)',
            (d['date'], d['owner'], d.get('envelope'), d.get('type'),
             d['amount'], d.get('notes'))
        )
        row = conn.execute('SELECT * FROM flux WHERE id=?', (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route('/api/flux/<int:fid>', methods=['PUT'])
@login_required
def update_flux(fid):
    d = request.json
    if not d or not validate_date(d.get('date')):
        return jsonify({'error': 'Date invalide'}), 400
    if not validate_number(d.get('amount'), allow_negative=True):
        return jsonify({'error': 'Montant invalide'}), 400
    with get_db() as conn:
        conn.execute(
            'UPDATE flux SET date=?, owner=?, envelope=?, type=?, amount=?, notes=? WHERE id=?',
            (d['date'], d['owner'], d.get('envelope'), d.get('type'),
             d['amount'], d.get('notes'), fid)
        )
        row = conn.execute('SELECT * FROM flux WHERE id=?', (fid,)).fetchone()
    return jsonify(dict(row))


@app.route('/api/flux/<int:fid>', methods=['DELETE'])
@login_required
def delete_flux(fid):
    with get_db() as conn:
        conn.execute('DELETE FROM flux WHERE id=?', (fid,))
    return '', 204


# — Synthèse —

@app.route('/api/synthese')
@login_required
def get_synthese():
    date = request.args.get('date')
    with get_db() as conn:
        if not date:
            row = conn.execute('SELECT MAX(date) as d FROM positions').fetchone()
            date = row['d']
        if not date:
            return jsonify({'date': None})
        rows       = conn.execute('SELECT * FROM positions WHERE date=?', (date,)).fetchall()
        entity_map = get_entity_map(conn, date)
        ref        = load_referential(conn)
        linked     = conn.execute(
            '''SELECT entity,
                      SUM(ownership_pct) as total_own,
                      SUM(debt_pct)      as total_debt
               FROM positions WHERE date=? AND entity IS NOT NULL
               GROUP BY entity''', (date,)
        ).fetchall()

    positions = [compute_position(dict(r), entity_map, ref) for r in rows]

    totals_by_owner = {}
    for owner in ref['owners']:
        ops = [p for p in positions if p['owner'] == owner]
        totals_by_owner[owner] = {
            'gross':       sum(p['gross_attributed'] for p in ops),
            'debt':        sum(p['debt_attributed'] for p in ops),
            'net':         sum(p['net_attributed'] for p in ops),
            'mobilizable': sum(p['mobilizable_value'] for p in ops),
        }

    totals_by_category = {}
    for cat in ref['categories']:
        ops = [p for p in positions if p['category'] == cat]
        if ops:
            totals_by_category[cat] = {
                'net':      sum(p['net_attributed'] for p in ops),
                'by_owner': {o: sum(p['net_attributed'] for p in ops if p['owner'] == o)
                             for o in OWNERS},
            }

    mobilizable_by_liquidity = {
        liq: sum(p['mobilizable_value'] for p in positions if p['liquidity'] == liq)
        for liq in ref['liquidity_order']
    }

    family = {
        'gross': sum(t['gross'] for t in totals_by_owner.values()),
        'debt':  sum(t['debt']  for t in totals_by_owner.values()),
        'net':   sum(t['net']   for t in totals_by_owner.values()),
    }

    entity_warnings = []
    for r in linked:
        own  = r['total_own']  or 0
        debt = r['total_debt'] or 0
        if own > 1.02:
            entity_warnings.append({'entity': r['entity'], 'total_pct': round(own * 100),  'type': 'ownership'})
        if debt > 1.02:
            entity_warnings.append({'entity': r['entity'], 'total_pct': round(debt * 100), 'type': 'debt'})

    # ── Variation par rapport au snapshot précédent ──
    variation = None
    with get_db() as conn2:
        prev_row = conn2.execute(
            'SELECT DISTINCT date FROM positions WHERE date < ? ORDER BY date DESC LIMIT 1',
            (date,)
        ).fetchone()
    if prev_row:
        prev_date = prev_row['date']
        with get_db() as conn2:
            prev_rows      = conn2.execute('SELECT * FROM positions WHERE date=?', (prev_date,)).fetchall()
            prev_entity_map = get_entity_map(conn2, prev_date)
            prev_ref        = load_referential(conn2)
        prev_positions = [compute_position(dict(r), prev_entity_map, prev_ref) for r in prev_rows]
        prev_net   = sum(p['net_attributed'] for p in prev_positions)
        prev_gross = sum(p['gross_attributed'] for p in prev_positions)
        prev_debt  = sum(p['debt_attributed'] for p in prev_positions)
        prev_mob   = sum(p['mobilizable_value'] for p in prev_positions)
        variation = {
            'prev_date':  prev_date,
            'net_delta':   family['net'] - prev_net,
            'net_pct':     ((family['net'] - prev_net) / abs(prev_net) * 100) if prev_net != 0 else None,
            'gross_delta': family['gross'] - prev_gross,
            'debt_delta':  family['debt'] - prev_debt,
            'mob_delta':   sum(t['mobilizable'] for t in totals_by_owner.values()) - prev_mob,
        }

    return jsonify({
        'date':                    date,
        'family':                  family,
        'totals_by_owner':         totals_by_owner,
        'totals_by_category':      totals_by_category,
        'mobilizable_by_liquidity': mobilizable_by_liquidity,
        'entity_warnings':         entity_warnings,
        'variation':               variation,
    })


@app.route('/api/historique')
@login_required
def get_historique():
    group_by = request.args.get('group_by')  # 'envelope' | 'category' | None
    owner    = request.args.get('owner')
    with get_db() as conn:
        dates = [r['date'] for r in conn.execute(
            'SELECT DISTINCT date FROM positions ORDER BY date'
        ).fetchall()]
        ref = load_referential(conn)
        history = []
        for date in dates:
            rows       = conn.execute('SELECT * FROM positions WHERE date=?', (date,)).fetchall()
            entity_map = get_entity_map(conn, date)
            positions  = [compute_position(dict(r), entity_map, ref) for r in rows]
            if owner:
                positions = [p for p in positions if p['owner'] == owner]
            entry = {
                'date':       date,
                'family_net': sum(p['net_attributed'] for p in positions),
                'by_owner':   {o: sum(p['net_attributed'] for p in positions if p['owner'] == o)
                              for o in ref['owners']},
            }
            if group_by == 'envelope':
                by_env = {}
                for p in positions:
                    k = p.get('envelope') or 'Autre'
                    by_env[k] = by_env.get(k, 0) + (p['net_attributed'] or 0)
                entry['by_group'] = by_env
            elif group_by == 'category':
                by_cat = {}
                for p in positions:
                    k = p.get('category') or 'Autre'
                    by_cat[k] = by_cat.get(k, 0) + (p['net_attributed'] or 0)
                entry['by_group'] = by_cat
            history.append(entry)
    return jsonify(history)


# — Entités —

@app.route('/api/entities', methods=['GET'])
@login_required
def get_entities():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM entities ORDER BY name').fetchall()
    result = []
    for r in rows:
        e = dict(r)
        e['net_assets'] = (e['gross_assets'] or 0) - (e['debt'] or 0)
        result.append(e)
    return jsonify(result)


@app.route('/api/entities', methods=['POST'])
@login_required
def add_entity():
    d = request.json
    if not d or not d.get('name') or not validate_string(d.get('name'), 200):
        return jsonify({'error': 'Nom requis'}), 400
    if not validate_number(d.get('gross_assets')) or not validate_number(d.get('debt')):
        return jsonify({'error': 'Valeurs numériques invalides'}), 400
    today = datetime.now().strftime('%Y-%m-%d')
    gross = d.get('gross_assets', 0)
    debt  = d.get('debt', 0)
    with get_db() as conn:
        cur = conn.execute(
            '''INSERT INTO entities (name, type, valuation_mode, gross_assets, debt, comment)
               VALUES (?,?,?,?,?,?)''',
            (d['name'], d.get('type'), d.get('valuation_mode'), gross, debt, d.get('comment'))
        )
        conn.execute(
            '''INSERT OR REPLACE INTO entity_snapshots (entity_name, date, gross_assets, debt)
               VALUES (?,?,?,?)''',
            (d['name'], today, gross, debt)
        )
        row = conn.execute('SELECT * FROM entities WHERE id=?', (cur.lastrowid,)).fetchone()
    e = dict(row)
    e['net_assets'] = (e['gross_assets'] or 0) - (e['debt'] or 0)
    e['snapshot_date'] = today
    return jsonify(e), 201


@app.route('/api/entities/<int:eid>', methods=['PUT'])
@login_required
def update_entity(eid):
    d = request.json
    if not d or not d.get('name'):
        return jsonify({'error': 'Nom requis'}), 400
    if not validate_number(d.get('gross_assets')) or not validate_number(d.get('debt')):
        return jsonify({'error': 'Valeurs numériques invalides'}), 400
    today = datetime.now().strftime('%Y-%m-%d')
    gross = d.get('gross_assets', 0)
    debt  = d.get('debt', 0)
    with get_db() as conn:
        conn.execute(
            '''UPDATE entities SET name=?, type=?, valuation_mode=?,
               gross_assets=?, debt=?, comment=? WHERE id=?''',
            (d['name'], d.get('type'), d.get('valuation_mode'), gross, debt, d.get('comment'), eid)
        )
        conn.execute(
            '''INSERT OR REPLACE INTO entity_snapshots (entity_name, date, gross_assets, debt)
               VALUES (?,?,?,?)''',
            (d['name'], today, gross, debt)
        )
        row = conn.execute('SELECT * FROM entities WHERE id=?', (eid,)).fetchone()
    e = dict(row)
    e['net_assets'] = (e['gross_assets'] or 0) - (e['debt'] or 0)
    e['snapshot_date'] = today
    return jsonify(e)


@app.route('/api/entities/<int:eid>', methods=['DELETE'])
@login_required
def delete_entity(eid):
    with get_db() as conn:
        name = conn.execute('SELECT name FROM entities WHERE id=?', (eid,)).fetchone()
        if name:
            conn.execute('DELETE FROM entity_snapshots WHERE entity_name=?', (name['name'],))
        conn.execute('DELETE FROM entities WHERE id=?', (eid,))
    return '', 204


# — Import Excel —

@app.route('/api/import', methods=['POST'])
@login_required
def import_xlsx():
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Seuls les fichiers .xlsx sont acceptés'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        imported = 0

        with get_db() as conn:
            # Positions
            ws = wb['Positions']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 7:
                    continue
                date_val, owner, category = row[0], row[1], row[2]
                envelope, establishment   = row[3], row[4]
                value, debt               = row[5], row[6]
                notes                     = row[8] if len(row) > 8 else None
                entity                    = row[9] if len(row) > 9 else None
                ownership_pct_raw = row[10] if len(row) > 10 else None
                debt_pct_raw      = row[11] if len(row) > 11 else None

                if not date_val or not owner:
                    continue
                if value is None and debt is None and envelope is None and not entity:
                    continue  # ligne vide du template

                if entity:
                    value = 0
                    debt  = 0
                    ownership_pct = ownership_pct_raw if ownership_pct_raw is not None else 1.0
                    debt_pct = debt_pct_raw if debt_pct_raw is not None else ownership_pct
                else:
                    ownership_pct = ownership_pct_raw if ownership_pct_raw is not None else 1.0
                    debt_pct      = debt_pct_raw      if debt_pct_raw      is not None else 1.0

                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)[:10]

                # Déduplication : ignorer si la combinaison (date, owner, category, envelope, entity) existe déjà
                existing = conn.execute(
                    '''SELECT id FROM positions
                       WHERE date=? AND owner=? AND category=?
                         AND COALESCE(envelope,'')=? AND COALESCE(entity,'')=?''',
                    (date_str, owner, category or '', envelope or '', entity or '')
                ).fetchone()
                if existing:
                    continue

                conn.execute(
                    '''INSERT INTO positions
                       (date, owner, category, envelope, establishment,
                        value, debt, notes, entity, ownership_pct, debt_pct)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                    (date_str, owner, category or '',
                     envelope, establishment,
                     value or 0, debt or 0,
                     notes, entity,
                     ownership_pct, debt_pct)
                )
                imported += 1

            # Flux (si la feuille contient des données)
            if 'Flux' in wb.sheetnames:
                wf = wb['Flux']
                flux_imported = 0
                for row in wf.iter_rows(min_row=2, values_only=True):
                    if len(row) < 5:
                        continue
                    date_val, owner, envelope, ftype, amount = row[0], row[1], row[2], row[3], row[4]
                    notes = row[5] if len(row) > 5 else None
                    if not date_val or not owner or amount is None:
                        continue
                    if isinstance(date_val, datetime):
                        date_str = date_val.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date_val)[:10]
                    conn.execute(
                        'INSERT INTO flux (date, owner, envelope, type, amount, notes) VALUES (?,?,?,?,?,?)',
                        (date_str, owner, envelope, ftype, amount, notes)
                    )
                    flux_imported += 1

            # Entités
            entities_imported = 0
            if 'Entites' in wb.sheetnames:
                we = wb['Entites']
                for row in we.iter_rows(min_row=2, values_only=True):
                    if len(row) < 2:
                        continue
                    name = row[0]
                    if not name:
                        continue
                    etype          = row[1] if len(row) > 1 else None
                    valuation_mode = row[2] if len(row) > 2 else None
                    gross_assets   = row[3] if len(row) > 3 else 0
                    debt           = row[4] if len(row) > 4 else 0
                    comment        = row[6] if len(row) > 6 else None  # col F = actif net (calculé), col G = commentaire
                    # Upsert : update si le nom existe déjà
                    existing = conn.execute(
                        'SELECT id FROM entities WHERE name=?', (name,)
                    ).fetchone()
                    if existing:
                        conn.execute(
                            '''UPDATE entities SET type=?, valuation_mode=?,
                               gross_assets=?, debt=?, comment=? WHERE name=?''',
                            (etype, valuation_mode, gross_assets or 0, debt or 0, comment, name)
                        )
                    else:
                        conn.execute(
                            '''INSERT INTO entities (name, type, valuation_mode, gross_assets, debt, comment)
                               VALUES (?,?,?,?,?,?)''',
                            (name, etype, valuation_mode, gross_assets or 0, debt or 0, comment)
                        )
                    # snapshot à la date d'aujourd'hui
                    today = datetime.now().strftime('%Y-%m-%d')
                    conn.execute(
                        '''INSERT OR REPLACE INTO entity_snapshots (entity_name, date, gross_assets, debt)
                           VALUES (?,?,?,?)''',
                        (name, today, gross_assets or 0, debt or 0)
                    )
                    entities_imported += 1

        return jsonify({'imported': imported, 'entities': entities_imported})

    except Exception as e:
        return jsonify({'error': str(e)}), 400


# — Snapshot update —

@app.route('/api/positions/<int:pid>/snapshot-update', methods=['POST'])
@login_required
def snapshot_update(pid):
    """
    Crée un nouveau snapshot à target_date en copiant toutes les positions
    de source_date, sauf pid qui est remplacé par les nouvelles valeurs.
    Si target_date a déjà des positions, on les écrase.
    """
    d           = request.json
    source_date = d.get('source_date')
    target_date = d.get('target_date')
    new_values  = d.get('position')

    if not source_date or not target_date or not new_values:
        return jsonify({'error': 'source_date, target_date et position requis'}), 400

    with get_db() as conn:
        entity_map = get_entity_map(conn, target_date)
        ref        = load_referential(conn)

        # Récupère toutes les positions du snapshot source
        source_rows = conn.execute(
            'SELECT * FROM positions WHERE date=?', (source_date,)
        ).fetchall()

        if not source_rows:
            return jsonify({'error': f'Aucune position à la date {source_date}'}), 404

        # Si le snapshot cible existe déjà, on le supprime pour repartir propre
        conn.execute('DELETE FROM positions WHERE date=?', (target_date,))

        created = []
        for row in source_rows:
            r = dict(row)
            if r['id'] == pid:
                # Position modifiée : on prend les nouvelles valeurs
                entity = new_values.get('entity')
                stored_value = 0 if entity else new_values.get('value', 0)
                stored_debt  = 0 if entity else new_values.get('debt', 0)
                cur = conn.execute(
                    '''INSERT INTO positions
                       (date, owner, category, envelope, establishment,
                        value, debt, notes, entity, ownership_pct, debt_pct)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                    (target_date,
                     new_values['owner'], new_values['category'],
                     new_values.get('envelope'), new_values.get('establishment'),
                     stored_value, stored_debt,
                     new_values.get('notes'), entity,
                     new_values.get('ownership_pct', 1.0),
                     new_values.get('debt_pct', 1.0))
                )
            else:
                # Position inchangée : copie telle quelle
                entity = r['entity']
                cur = conn.execute(
                    '''INSERT INTO positions
                       (date, owner, category, envelope, establishment,
                        value, debt, notes, entity, ownership_pct, debt_pct)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                    (target_date,
                     r['owner'], r['category'], r['envelope'], r['establishment'],
                     r['value'], r['debt'], r['notes'], r['entity'],
                     r['ownership_pct'], r['debt_pct'])
                )
            new_row = conn.execute('SELECT * FROM positions WHERE id=?', (cur.lastrowid,)).fetchone()
            created.append(compute_position(dict(new_row), entity_map, ref))

    return jsonify({'target_date': target_date, 'count': len(created)}), 201


# — Reset —

@app.route('/api/import-json', methods=['POST'])
@login_required
def import_json():
    data = request.json
    if not data:
        return jsonify({'error': 'Corps JSON manquant'}), 400

    imported = {'positions': 0, 'flux': 0, 'entities': 0, 'entity_snapshots': 0}

    with get_db() as conn:
        # Entités
        for e in data.get('entities', []):
            name = e.get('name')
            if not name:
                continue
            existing = conn.execute('SELECT id FROM entities WHERE name=?', (name,)).fetchone()
            if existing:
                conn.execute(
                    'UPDATE entities SET type=?, valuation_mode=?, gross_assets=?, debt=?, comment=? WHERE name=?',
                    (e.get('type'), e.get('valuation_mode'), e.get('gross_assets', 0), e.get('debt', 0), e.get('comment'), name)
                )
            else:
                conn.execute(
                    'INSERT INTO entities (name, type, valuation_mode, gross_assets, debt, comment) VALUES (?,?,?,?,?,?)',
                    (name, e.get('type'), e.get('valuation_mode'), e.get('gross_assets', 0), e.get('debt', 0), e.get('comment'))
                )
            imported['entities'] += 1

        # Snapshots entités
        for s in data.get('entity_snapshots', []):
            if not s.get('entity_name') or not s.get('date'):
                continue
            conn.execute(
                'INSERT OR REPLACE INTO entity_snapshots (entity_name, date, gross_assets, debt) VALUES (?,?,?,?)',
                (s['entity_name'], s['date'], s.get('gross_assets', 0), s.get('debt', 0))
            )
            imported['entity_snapshots'] += 1

        # Positions
        for p in data.get('positions', []):
            if not p.get('date') or not p.get('owner'):
                continue
            existing = conn.execute(
                '''SELECT id FROM positions WHERE date=? AND owner=? AND category=?
                   AND COALESCE(envelope,'')=? AND COALESCE(entity,'')=?''',
                (p['date'], p['owner'], p.get('category', ''), p.get('envelope') or '', p.get('entity') or '')
            ).fetchone()
            if existing:
                continue
            conn.execute(
                '''INSERT INTO positions (date, owner, category, envelope, establishment,
                   value, debt, notes, entity, ownership_pct, debt_pct) VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (p['date'], p['owner'], p.get('category', ''), p.get('envelope'), p.get('establishment'),
                 p.get('value', 0), p.get('debt', 0), p.get('notes'), p.get('entity'),
                 p.get('ownership_pct', 1.0), p.get('debt_pct', 1.0))
            )
            imported['positions'] += 1

        # Flux
        for f in data.get('flux', []):
            if not f.get('date') or not f.get('owner') or f.get('amount') is None:
                continue
            conn.execute(
                'INSERT INTO flux (date, owner, envelope, type, amount, notes) VALUES (?,?,?,?,?,?)',
                (f['date'], f['owner'], f.get('envelope'), f.get('type'), f['amount'], f.get('notes'))
            )
            imported['flux'] += 1

    return jsonify(imported)


@app.route('/api/entity-snapshots')
@login_required
def get_entity_snapshots():
    entity_name = request.args.get('entity')
    with get_db() as conn:
        if entity_name:
            rows = conn.execute(
                'SELECT * FROM entity_snapshots WHERE entity_name=? ORDER BY date DESC',
                (entity_name,)
            ).fetchall()
        else:
            rows = conn.execute(
                'SELECT * FROM entity_snapshots ORDER BY entity_name, date DESC'
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/entity-snapshots/<int:sid>', methods=['DELETE'])
@login_required
def delete_entity_snapshot(sid):
    with get_db() as conn:
        conn.execute('DELETE FROM entity_snapshots WHERE id=?', (sid,))
    return '', 204


@app.route('/api/reset', methods=['POST'])
@login_required
def reset_db():
    with get_db() as conn:
        conn.executescript(
            'DELETE FROM positions; DELETE FROM flux; '
            'DELETE FROM entities; DELETE FROM entity_snapshots;'
        )
    return jsonify({'ok': True})


# — Export JSON —

@app.route('/api/export')
@login_required
def export_data():
    with get_db() as conn:
        positions = [dict(r) for r in conn.execute(
            'SELECT * FROM positions ORDER BY date, owner'
        ).fetchall()]
        flux = [dict(r) for r in conn.execute(
            'SELECT * FROM flux ORDER BY date'
        ).fetchall()]
        entities = [dict(r) for r in conn.execute(
            'SELECT * FROM entities ORDER BY name'
        ).fetchall()]
        entity_snapshots = [dict(r) for r in conn.execute(
            'SELECT * FROM entity_snapshots ORDER BY entity_name, date'
        ).fetchall()]
    return jsonify({
        'positions': positions,
        'flux': flux,
        'entities': entities,
        'entity_snapshots': entity_snapshots,
    })


# — Référentiel —

# — Allocation cible (persistance DB) —

@app.route('/api/targets', methods=['GET'])
@login_required
def get_targets():
    with get_db() as conn:
        row = conn.execute("SELECT value FROM config WHERE key='allocation_targets'").fetchone()
    if row:
        try:
            return jsonify(json.loads(row['value']))
        except Exception:
            pass
    return jsonify({})


@app.route('/api/targets', methods=['PUT'])
@login_required
def save_targets():
    data = request.json
    if not isinstance(data, dict):
        return jsonify({'error': 'Objet JSON attendu'}), 400
    with get_db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO config (key, value) VALUES ('allocation_targets', ?)",
            (json.dumps(data),)
        )
    return jsonify({'ok': True})


# — Alertes (persistance DB) —

@app.route('/api/alerts', methods=['GET'])
@login_required
def get_alerts():
    with get_db() as conn:
        row = conn.execute("SELECT value FROM config WHERE key='user_alerts'").fetchone()
    if row:
        try:
            return jsonify(json.loads(row['value']))
        except Exception:
            pass
    return jsonify([])


@app.route('/api/alerts', methods=['PUT'])
@login_required
def save_alerts_api():
    data = request.json
    if not isinstance(data, list):
        return jsonify({'error': 'Tableau JSON attendu'}), 400
    with get_db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO config (key, value) VALUES ('user_alerts', ?)",
            (json.dumps(data),)
        )
    return jsonify({'ok': True})


@app.route('/api/referential', methods=['GET'])
@login_required
def get_referential_api():
    with get_db() as conn:
        ref = load_referential(conn)
    return jsonify(ref)


@app.route('/api/referential', methods=['PUT'])
@login_required
def save_referential():
    data = request.json
    required = ['owners', 'categories', 'category_mobilizable', 'envelope_meta']
    for k in required:
        if k not in data:
            return jsonify({'error': f'Champ manquant : {k}'}), 400
    if not data['owners']:
        return jsonify({'error': 'La liste des propriétaires ne peut pas être vide'}), 400
    # liquidity_order est fixe, on l'exclut du stockage
    data.pop('liquidity_order', None)
    with get_db() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO config (key, value) VALUES ('referential', ?)",
            (json.dumps(data),)
        )
    return jsonify({'ok': True})


# ─── Démarrage ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    print("✓ Base initialisée — http://localhost:5000")
    app.run(debug=True, port=5000)
